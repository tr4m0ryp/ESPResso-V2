#!/usr/bin/env python3
"""
Extract AWARE20 characterization factors from XLSX into CSV lookup tables.

Reads: data/extern/AWARE20_Countries_and_Regions.xlsx
Writes:
  - data/datasets/pre-model/final/water_footprint/aware_factors_agri.csv
  - data/datasets/pre-model/final/water_footprint/aware_factors_nonagri.csv
  - data/datasets/pre-model/final/water_footprint/aware_country_aliases.csv
"""

import csv
import os
import sys

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run:")
    print("  python3 -m venv /tmp/xlsx_env && /tmp/xlsx_env/bin/pip install openpyxl")
    sys.exit(1)

# -- Paths ---------------------------------------------------------------

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.abspath(os.path.join(SCRIPT_DIR, "..", "..", "..", ".."))
XLSX_PATH = os.path.join(PROJECT_ROOT, "data", "extern",
                         "AWARE20_Countries_and_Regions.xlsx")
OUTPUT_DIR = os.path.join(PROJECT_ROOT, "data", "datasets", "pre-model",
                          "final", "water_footprint")

# -- Country alias table (D6 spec) ---------------------------------------

COUNTRY_ALIASES = [
    ("Turkey", "Turkiye"),
    ("USA", "United States of America"),
    ("UK", "United Kingdom"),
    ("England", "United Kingdom"),
    ("Scotland", "United Kingdom"),
    ("Czech Republic", "Czechia"),
    ("Kitwe", "Zambia"),
    ("Lusaka", "Zambia"),
    ("Tashkent", "Uzbekistan"),
    ("Fergana Valley", "Uzbekistan"),
]


def extract_country_sheet(wb, sheet_name):
    """Extract rows from a CFs_agri or CFs_nonagri sheet.

    Returns list of dicts with keys:
      country_name, ecoinvent_shortname, iso3, aware_cf_annual
    Also returns count of skipped NotDefined rows.
    """
    ws = wb[sheet_name]
    rows_out = []
    skipped = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        # Column layout: ID(0), conform_GLAM(1), conform_ecoinvent(2),
        #   GLAM_country_name(3), ecoinvent_country_name(4), GLAM_ISO3(5),
        #   ecoinvent_shortname(6), ecoinvent_collection(7), Annual(8), ...
        ecoinvent_name = row[4]
        glam_name = row[3]
        iso3 = row[5]
        shortname = row[6]
        annual = row[8]

        # Skip NotDefined or missing annual values
        if annual is None or str(annual).strip() == "NotDefined":
            skipped += 1
            continue

        # Primary: ecoinvent_country_name; fallback: GLAM_country_name
        # Treat "NotDefined" same as empty/None for all text fields
        country_name = ecoinvent_name
        if (country_name is None
                or str(country_name).strip() in ("", "NotDefined")):
            country_name = glam_name
        if (country_name is None
                or str(country_name).strip() in ("", "NotDefined")):
            country_name = ""

        country_name = str(country_name).strip()
        shortname = str(shortname).strip() if shortname else ""
        iso3 = str(iso3).strip() if iso3 else ""
        # Clean "NotDefined" values in non-annual fields
        if iso3 == "NotDefined":
            iso3 = ""
        if shortname == "NotDefined":
            shortname = ""

        # Round to 3 decimal places
        try:
            annual_val = round(float(annual), 3)
        except (ValueError, TypeError):
            skipped += 1
            continue

        rows_out.append({
            "country_name": country_name,
            "ecoinvent_shortname": shortname,
            "iso3": iso3,
            "aware_cf_annual": annual_val,
        })

    return rows_out, skipped


def extract_regional_sheet(wb, sheet_name):
    """Extract rows from a CFs_add_agri or CFs_add_nonagri sheet.

    Returns list of dicts matching the same schema as country sheets.
    Regional rows use Region as country_name and ecoinvent_shortname,
    and iso3 is set equal to Region.
    """
    ws = wb[sheet_name]
    rows_out = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        # Column layout: RegionType(0), Region(1), Jan..Dec(2-13), annual(14)
        region = row[1]
        annual = row[14]  # Column O = index 14

        if region is None or annual is None:
            continue
        if str(annual).strip() == "NotDefined":
            continue

        region = str(region).strip()
        try:
            annual_val = round(float(annual), 3)
        except (ValueError, TypeError):
            continue

        rows_out.append({
            "country_name": region,
            "ecoinvent_shortname": region,
            "iso3": region,
            "aware_cf_annual": annual_val,
        })

    return rows_out


def write_factors_csv(rows, output_path):
    """Write factor rows to CSV with UTF-8 encoding."""
    fieldnames = ["country_name", "ecoinvent_shortname", "iso3",
                  "aware_cf_annual"]
    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def write_aliases_csv(output_path):
    """Write the static country alias table."""
    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["alias", "canonical_name"])
        for alias, canonical in COUNTRY_ALIASES:
            writer.writerow([alias, canonical])


def main():
    if not os.path.exists(XLSX_PATH):
        print(f"ERROR: XLSX not found at {XLSX_PATH}")
        sys.exit(1)

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)

    # -- Agricultural factors ---------------------------------------------
    agri_rows, agri_skipped = extract_country_sheet(wb, "CFs_agri")
    agri_regional = extract_regional_sheet(wb, "CFs_add_agri")
    agri_total_raw = len(agri_rows) + agri_skipped
    agri_all = agri_rows + agri_regional

    agri_path = os.path.join(OUTPUT_DIR, "aware_factors_agri.csv")
    write_factors_csv(agri_all, agri_path)

    # -- Non-agricultural factors -----------------------------------------
    nonagri_rows, nonagri_skipped = extract_country_sheet(wb, "CFs_nonagri")
    nonagri_regional = extract_regional_sheet(wb, "CFs_add_nonagri")
    nonagri_total_raw = len(nonagri_rows) + nonagri_skipped
    nonagri_all = nonagri_rows + nonagri_regional

    nonagri_path = os.path.join(OUTPUT_DIR, "aware_factors_nonagri.csv")
    write_factors_csv(nonagri_all, nonagri_path)

    # -- Country aliases --------------------------------------------------
    aliases_path = os.path.join(OUTPUT_DIR, "aware_country_aliases.csv")
    write_aliases_csv(aliases_path)

    wb.close()

    # -- Summary ----------------------------------------------------------
    print("AWARE extraction complete.")
    print()
    print(f"Agricultural (CFs_agri + CFs_add_agri):")
    print(f"  Country rows extracted: {len(agri_rows)}")
    print(f"  Country rows skipped (NotDefined Annual): {agri_skipped}"
          f" ({agri_skipped / agri_total_raw * 100:.1f}%)")
    if agri_skipped / agri_total_raw > 0.10:
        print(f"  WARNING: >10% of rows had NotDefined Annual values.")
    print(f"  Regional fallback rows: {len(agri_regional)}")
    print(f"  Total rows written: {len(agri_all)}")
    print()
    print(f"Non-agricultural (CFs_nonagri + CFs_add_nonagri):")
    print(f"  Country rows extracted: {len(nonagri_rows)}")
    print(f"  Country rows skipped (NotDefined Annual): {nonagri_skipped}"
          f" ({nonagri_skipped / nonagri_total_raw * 100:.1f}%)")
    if nonagri_skipped / nonagri_total_raw > 0.10:
        print(f"  WARNING: >10% of rows had NotDefined Annual values.")
    print(f"  Regional fallback rows: {len(nonagri_regional)}")
    print(f"  Total rows written: {len(nonagri_all)}")
    print()
    print(f"Country aliases: {len(COUNTRY_ALIASES)} rows")
    print()

    # Verify GLO row exists
    agri_glo = [r for r in agri_all if r["country_name"] == "GLO"]
    nonagri_glo = [r for r in nonagri_all if r["country_name"] == "GLO"]
    print(f"GLO row in agri: {'YES' if agri_glo else 'MISSING'}"
          f" (CF={agri_glo[0]['aware_cf_annual'] if agri_glo else 'N/A'})")
    print(f"GLO row in nonagri: {'YES' if nonagri_glo else 'MISSING'}"
          f" (CF={nonagri_glo[0]['aware_cf_annual'] if nonagri_glo else 'N/A'})")
    print()
    print("Output files:")
    print(f"  {agri_path}")
    print(f"  {nonagri_path}")
    print(f"  {aliases_path}")


if __name__ == "__main__":
    main()
